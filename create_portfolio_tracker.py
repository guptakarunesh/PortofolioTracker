from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import PieChart, LineChart, DoughnutChart, BarChart, Reference


OUT_FILE = "Indian_Investment_Portfolio_Tracker.xlsx"

wb = Workbook()

# Remove default sheet and create all required sheets in order
wb.remove(wb.active)
sheet_names = [
    "Dashboard",
    "Banking",
    "MutualFunds",
    "Stocks",
    "Gold",
    "RealEstate",
    "Retirement",
    "Insurance",
    "Loans",
    "OtherAssets",
    "TransactionHistory",
    "Reminders",
    "Settings",
    "Instructions",
]
for n in sheet_names:
    wb.create_sheet(title=n)

# Shared styles
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF")
total_fill = PatternFill("solid", fgColor="FFC000")
total_font = Font(bold=True, color="000000")
thin = Side(style="thin", color="000000")
all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center")

currency_fmt = "₹#,##0"
percent_fmt = '0.00"%"'
date_fmt = "DD-MMM-YYYY"


def style_header(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = all_border


def style_total_row(ws, row, start_col, end_col):
    for c in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = total_fill
        cell.font = total_font
        cell.border = all_border


def border_range(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = all_border


# ---------------- Sheet 1: Dashboard ----------------
ws = wb["Dashboard"]
ws.merge_cells("A1:D1")
ws["A1"] = "INDIAN INVESTMENT PORTFOLIO TRACKER"
ws["A1"].font = Font(bold=True, size=16, color="0070C0")
ws["A1"].alignment = Alignment(horizontal="center")

ws["A2"] = "Last Updated"
ws["B2"] = date.today()
ws["B2"].number_format = date_fmt

ws["A4"] = "Net Worth Summary"
ws["A4"].font = Font(bold=True)

ws["A5"] = "Total Assets"
ws["B5"] = "=Banking!D11+MutualFunds!G21+Stocks!G21+Gold!G7+RealEstate!F11+Retirement!C11+Insurance!I11+OtherAssets!E11"
ws["A6"] = "Total Liabilities"
ws["B6"] = "=Loans!E11"
ws["A7"] = "NET WORTH"
ws["B7"] = "=B5-B6"
ws["A7"].font = Font(bold=True, size=14)
ws["B7"].font = Font(bold=True, size=14, color="008000")

for c in ["B5", "B6", "B7"]:
    ws[c].number_format = currency_fmt

ws["A9"] = "Asset Allocation"
ws["A9"].font = Font(bold=True)

headers = ["Category", "Current Value (₹)", "% of Total", "Notes"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=10, column=i, value=h)
style_header(ws, 10, 1, 4)

rows = [
    ("Banking & Deposits", "=Banking!D11", "Primary liquid assets"),
    ("Market Investments", "=MutualFunds!G21+Stocks!G21", "MF + Direct Equity"),
    ("Precious Metals", "=Gold!G7", "Gold and silver"),
    ("Real Estate", "=RealEstate!F11", "Property market value"),
    ("Retirement Funds", "=Retirement!C11", "EPF/PPF/NPS"),
    ("Insurance (Cash Value)", "=Insurance!I11", "Surrender value only"),
    ("Other Assets", "=OtherAssets!E11", "Vehicles, crypto, ESOPs"),
]
for idx, (cat, valf, note) in enumerate(rows, start=11):
    ws.cell(row=idx, column=1, value=cat)
    ws.cell(row=idx, column=2, value=valf)
    ws.cell(row=idx, column=3, value=f"=B{idx}/$B$18*100")
    ws.cell(row=idx, column=4, value=note)
    ws.cell(row=idx, column=2).number_format = currency_fmt
    ws.cell(row=idx, column=3).number_format = percent_fmt

ws["A18"] = "TOTAL ASSETS"
ws["B18"] = "=SUM(B11:B17)"
ws["C18"] = "=B18/$B$18*100"
ws["B18"].number_format = currency_fmt
ws["C18"].number_format = percent_fmt
style_total_row(ws, 18, 1, 4)
border_range(ws, 10, 18, 1, 4)

for col, w in {"A": 30, "B": 20, "C": 12, "D": 32}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 2: Banking ----------------
ws = wb["Banking"]
headers = [
    "Account Type", "Bank Name", "Account Number", "Current Balance (₹)",
    "Interest Rate (%)", "Maturity Date", "Remarks"
]
ws.append(headers)
style_header(ws, 1, 1, 7)

bank_data = [
    ["Savings", "HDFC Bank", "XXXX1234", 50000, 3.5, None, "Primary Account"],
    ["FD", "SBI", "XXXX5678", 200000, 7.1, datetime(2026, 8, 15), "5 Year FD"],
    ["RD", "ICICI", "XXXX9012", 50000, 6.8, datetime(2026, 12, 30), "Monthly ₹5000"],
    ["PPF", "SBI", "XXXX3456", 500000, 7.1, datetime(2041, 3, 31), "Annual Contribution"],
]
for r in bank_data:
    ws.append(r)

for r in range(2, 11):
    ws.cell(r, 4).number_format = currency_fmt
    ws.cell(r, 5).number_format = percent_fmt
    ws.cell(r, 6).number_format = date_fmt

ws["A11"] = "TOTAL"
ws["D11"] = "=SUM(D2:D10)"
ws["A11"].font = Font(bold=True)
ws["D11"].number_format = currency_fmt
style_total_row(ws, 11, 1, 7)
border_range(ws, 1, 11, 1, 7)

for col, w in {"A": 15, "B": 15, "C": 15, "D": 18, "E": 12, "F": 15, "G": 20}.items():
    ws.column_dimensions[col].width = w

# Data validation
bank_dv = DataValidation(type="list", formula1='"Savings,Current,FD,RD,PPF,NSC,KVP,SCSS"')
ws.add_data_validation(bank_dv)
bank_dv.add("A2:A100")


# ---------------- Sheet 3: MutualFunds ----------------
ws = wb["MutualFunds"]
headers = [
    "Folio Number", "AMC", "Scheme Name", "Category", "Units", "Current NAV",
    "Current Value (₹)", "Invested Amount (₹)", "Gain/Loss (₹)", "Gain/Loss (%)",
    "XIRR (%)", "SIP Amount (₹)", "SIP Date"
]
ws.append(headers)
style_header(ws, 1, 1, 13)

mf_data = [
    ["12345/67", "HDFC MF", "HDFC Flexi Cap", "Equity", 1000.50, 850.25, "=E2*F2", 750000, "=G2-H2", "=I2/H2", 15.5, 10000, "5th"],
    ["23456/78", "SBI MF", "SBI Bluechip", "Equity", 500, 95.50, "=E3*F3", 45000, "=G3-H3", "=I3/H3", 12.3, 5000, "10th"],
]
for r in mf_data:
    ws.append(r)

for r in range(2, 21):
    for c in [7, 8, 9, 12]:
        ws.cell(r, c).number_format = currency_fmt
    for c in [10, 11]:
        ws.cell(r, c).number_format = percent_fmt

ws["A21"] = "TOTAL"
ws["G21"] = "=SUM(G2:G20)"
ws["H21"] = "=SUM(H2:H20)"
ws["I21"] = "=SUM(I2:I20)"
ws["J21"] = "=I21/H21"
for c in ["G21", "H21", "I21"]:
    ws[c].number_format = currency_fmt
ws["J21"].number_format = percent_fmt
style_total_row(ws, 21, 1, 13)
border_range(ws, 1, 21, 1, 13)

for col, w in {
    "A": 14, "B": 14, "C": 22, "D": 12, "E": 10, "F": 12,
    "G": 16, "H": 16, "I": 14, "J": 12, "K": 10, "L": 12, "M": 10
}.items():
    ws.column_dimensions[col].width = w

mf_dv = DataValidation(type="list", formula1='"Equity,Debt,Hybrid,Liquid,ELSS"')
ws.add_data_validation(mf_dv)
mf_dv.add("D2:D100")


# ---------------- Sheet 4: Stocks ----------------
ws = wb["Stocks"]
headers = [
    "Stock Symbol", "Company Name", "Exchange", "Quantity", "Avg Buy Price (₹)",
    "Current Price (₹)", "Current Value (₹)", "Invested Amount (₹)",
    "Unrealized Gain/Loss (₹)", "Unrealized Gain/Loss (%)", "Demat Account"
]
ws.append(headers)
style_header(ws, 1, 1, 11)

stocks_data = [
    ["RELIANCE", "Reliance Industries", "NSE", 50, 2400, 2850, "=D2*F2", "=D2*E2", "=G2-H2", "=I2/H2", "Zerodha"],
    ["TCS", "Tata Consultancy", "NSE", 25, 3500, 4100, "=D3*F3", "=D3*E3", "=G3-H3", "=I3/H3", "Zerodha"],
]
for r in stocks_data:
    ws.append(r)

for r in range(2, 21):
    for c in [5, 6, 7, 8, 9]:
        ws.cell(r, c).number_format = currency_fmt
    ws.cell(r, 10).number_format = percent_fmt

ws["A21"] = "TOTAL"
ws["G21"] = "=SUM(G2:G20)"
ws["H21"] = "=SUM(H2:H20)"
ws["I21"] = "=SUM(I2:I20)"
ws["J21"] = "=I21/H21"
for c in ["G21", "H21", "I21"]:
    ws[c].number_format = currency_fmt
ws["J21"].number_format = percent_fmt
style_total_row(ws, 21, 1, 11)
border_range(ws, 1, 21, 1, 11)

for col, w in {
    "A": 14, "B": 20, "C": 10, "D": 10, "E": 14, "F": 14,
    "G": 14, "H": 14, "I": 16, "J": 14, "K": 14
}.items():
    ws.column_dimensions[col].width = w

stocks_dv = DataValidation(type="list", formula1='"NSE,BSE"')
ws.add_data_validation(stocks_dv)
stocks_dv.add("C2:C100")


# ---------------- Sheet 5: Gold ----------------
ws = wb["Gold"]
headers = [
    "Type", "Sub-Type", "Weight (grams)", "Purchase Price/gram (₹)", "Purchase Date",
    "Current Rate/gram (₹)", "Current Value (₹)", "Gain/Loss (₹)", "Location/Details"
]
ws.append(headers)
style_header(ws, 1, 1, 9)

gold_data = [
    ["Gold", "Physical Jewelry", 50, 4500, datetime(2020, 1, 1), 6500, "=C2*F2", "=G2-(C2*D2)", "Home Locker"],
    ["Gold", "Coins (24K)", 20, 5200, datetime(2022, 8, 15), 6500, "=C3*F3", "=G3-(C3*D3)", "Bank Locker"],
    ["Gold", "Digital Gold", 10, 5500, datetime(2024, 1, 10), 6500, "=C4*F4", "=G4-(C4*D4)", "Paytm"],
    ["Gold", "Sovereign Gold Bonds", 16, 5000, datetime(2023, 4, 1), 6500, "=C5*F5", "=G5-(C5*D5)", "Demat"],
    ["Silver", "Physical Bars", 500, 65, datetime(2021, 6, 1), 75, "=C6*F6", "=G6-(C6*D6)", "Home"],
]
for r in gold_data:
    ws.append(r)

for r in range(2, 8):
    ws.cell(r, 4).number_format = currency_fmt
    ws.cell(r, 5).number_format = date_fmt
    ws.cell(r, 6).number_format = currency_fmt
    ws.cell(r, 7).number_format = currency_fmt
    ws.cell(r, 8).number_format = currency_fmt

ws["A7"] = "TOTAL"
ws["C7"] = "=SUM(C2:C6)"
ws["G7"] = "=SUM(G2:G6)"
ws["H7"] = "=SUM(H2:H6)"
ws["G7"].number_format = currency_fmt
ws["H7"].number_format = currency_fmt
style_total_row(ws, 7, 1, 9)
border_range(ws, 1, 7, 1, 9)

for col, w in {"A": 10, "B": 22, "C": 12, "D": 18, "E": 14, "F": 18, "G": 16, "H": 14, "I": 20}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 6: RealEstate ----------------
ws = wb["RealEstate"]
headers = [
    "Property Type", "Address", "City", "Purchase Date", "Purchase Price (₹)",
    "Current Market Value (₹)", "Loan Outstanding (₹)", "Net Equity (₹)",
    "Rental Income (₹/month)", "Remarks"
]
ws.append(headers)
style_header(ws, 1, 1, 10)

re_data = [
    ["Residential", "Sector 62, Noida", "Noida", datetime(2020, 3, 15), 7500000, 9500000, 3000000, "=F2-G2", 25000, "Self-occupied"],
    ["Plot", "Greater Noida", "Noida", datetime(2018, 1, 1), 2500000, 4000000, 0, "=F3-G3", None, "Vacant"],
]
for r in re_data:
    ws.append(r)

for r in range(2, 11):
    ws.cell(r, 4).number_format = date_fmt
    for c in [5, 6, 7, 8, 9]:
        ws.cell(r, c).number_format = currency_fmt

ws["A11"] = "TOTAL"
ws["E11"] = "=SUM(E2:E10)"
ws["F11"] = "=SUM(F2:F10)"
ws["G11"] = "=SUM(G2:G10)"
ws["H11"] = "=SUM(H2:H10)"
ws["I11"] = "=SUM(I2:I10)"
for c in ["E11", "F11", "G11", "H11", "I11"]:
    ws[c].number_format = currency_fmt
style_total_row(ws, 11, 1, 10)
border_range(ws, 1, 11, 1, 10)

for col, w in {
    "A": 16, "B": 24, "C": 12, "D": 14, "E": 16,
    "F": 18, "G": 16, "H": 14, "I": 18, "J": 18
}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 7: Retirement ----------------
ws = wb["Retirement"]
headers = [
    "Account Type", "Account Number", "Current Balance (₹)", "Employer Contribution (₹)",
    "Employee Contribution (₹)", "Interest Rate (%)", "Maturity/Retirement Date", "Remarks"
]
ws.append(headers)
style_header(ws, 1, 1, 8)

ret_data = [
    ["EPF", "PF/XX/12345", 800000, "1800/month", "1800/month", "8.25%", datetime(2050, 3, 31), "Auto-deducted"],
    ["PPF", "PPF123456", 500000, "-", "150000/year", "7.1%", datetime(2041, 3, 31), "15-year lock-in"],
    ["NPS Tier-I", "PRAN123456", 300000, "50000/year", "50000/year", "Market-linked", datetime(2050, 3, 31), "Tax benefit"],
]
for r in ret_data:
    ws.append(r)

for r in range(2, 11):
    ws.cell(r, 3).number_format = currency_fmt
    ws.cell(r, 7).number_format = date_fmt

ws["A11"] = "TOTAL"
ws["C11"] = "=SUM(C2:C10)"
ws["C11"].number_format = currency_fmt
style_total_row(ws, 11, 1, 8)
border_range(ws, 1, 11, 1, 8)

for col, w in {"A": 14, "B": 16, "C": 16, "D": 18, "E": 18, "F": 14, "G": 18, "H": 18}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 8: Insurance ----------------
ws = wb["Insurance"]
headers = [
    "Policy Type", "Insurer", "Policy Number", "Sum Assured (₹)", "Premium Amount (₹)",
    "Premium Frequency", "Next Due Date", "Maturity Date", "Cash/Surrender Value (₹)",
    "Nominee", "Status"
]
ws.append(headers)
style_header(ws, 1, 1, 11)

ins_data = [
    ["Term Life", "HDFC Life", "TERM123456", 10000000, 15000, "Annual", datetime(2026, 4, 1), datetime(2050, 4, 1), 0, "Spouse", "Active"],
    ["Endowment", "LIC", "LIC987654", 1000000, 50000, "Annual", datetime(2026, 5, 15), datetime(2035, 5, 15), 350000, "Spouse", "Active"],
    ["Health", "Star Health", "HEALTH5678", 1000000, 18000, "Annual", datetime(2026, 6, 10), "Annual renewal", 0, "Self", "Active"],
    ["Vehicle", "ICICI Lombard", "VEH123", 800000, 12000, "Annual", datetime(2026, 7, 20), "Annual renewal", 0, "Self", "Active"],
]
for r in ins_data:
    ws.append(r)

for r in range(2, 11):
    for c in [4, 5, 9]:
        ws.cell(r, c).number_format = currency_fmt
    ws.cell(r, 7).number_format = date_fmt
    if isinstance(ws.cell(r, 8).value, (datetime, date)):
        ws.cell(r, 8).number_format = date_fmt

ws["A11"] = "TOTAL"
ws["D11"] = "=SUM(D2:D10)"
ws["E11"] = "=SUM(E2:E10)"
ws["I11"] = "=SUM(I2:I10)"
for c in ["D11", "E11", "I11"]:
    ws[c].number_format = currency_fmt
style_total_row(ws, 11, 1, 11)
border_range(ws, 1, 11, 1, 11)

for col, w in {
    "A": 14, "B": 14, "C": 14, "D": 16, "E": 16, "F": 16,
    "G": 14, "H": 16, "I": 18, "J": 12, "K": 10
}.items():
    ws.column_dimensions[col].width = w

ins_type_dv = DataValidation(type="list", formula1='"Term Life,Endowment,ULIP,Health,Vehicle,Property"')
ws.add_data_validation(ins_type_dv)
ins_type_dv.add("A2:A100")

ins_freq_dv = DataValidation(type="list", formula1='"Monthly,Quarterly,Half-yearly,Annual"')
ws.add_data_validation(ins_freq_dv)
ins_freq_dv.add("F2:F100")


# ---------------- Sheet 9: Loans ----------------
ws = wb["Loans"]
headers = [
    "Loan Type", "Lender", "Loan Account", "Original Amount (₹)", "Outstanding Amount (₹)",
    "Interest Rate (%)", "EMI Amount (₹)", "EMI Date", "Tenure Remaining", "End Date"
]
ws.append(headers)
style_header(ws, 1, 1, 10)

loan_data = [
    ["Home Loan", "HDFC Bank", "HL123456", 5000000, 3000000, 8.5, 45000, "5th", "120 months", datetime(2036, 2, 5)],
    ["Car Loan", "ICICI Bank", "CL789012", 800000, 350000, 9.2, 18000, "10th", "24 months", datetime(2028, 2, 10)],
    ["Personal Loan", "SBI", "PL345678", 200000, 120000, 11.5, 12000, "15th", "12 months", datetime(2027, 2, 15)],
]
for r in loan_data:
    ws.append(r)

for r in range(2, 11):
    for c in [4, 5, 7]:
        ws.cell(r, c).number_format = currency_fmt
    ws.cell(r, 6).number_format = percent_fmt
    ws.cell(r, 10).number_format = date_fmt

ws["A11"] = "TOTAL"
ws["D11"] = "=SUM(D2:D10)"
ws["E11"] = "=SUM(E2:E10)"
ws["G11"] = "=SUM(G2:G10)"
for c in ["D11", "E11", "G11"]:
    ws[c].number_format = currency_fmt
style_total_row(ws, 11, 1, 10)
border_range(ws, 1, 11, 1, 10)

for col, w in {"A": 14, "B": 14, "C": 14, "D": 16, "E": 18, "F": 12, "G": 14, "H": 10, "I": 16, "J": 14}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 10: OtherAssets ----------------
ws = wb["OtherAssets"]
headers = [
    "Asset Type", "Description", "Purchase Date", "Purchase Value (₹)",
    "Current Value (₹)", "Depreciation Rate (%)", "Remarks"
]
ws.append(headers)
style_header(ws, 1, 1, 7)

other_data = [
    ["Vehicle", "Honda City 2022", datetime(2022, 1, 15), 1200000, 800000, "15%/year", "Registration: DL-XX-XXXX"],
    ["Cryptocurrency", "Bitcoin", "Various", 100000, 150000, "Volatile", "WazirX account"],
    ["ESOPs", "Company Stock", datetime(2024, 1, 1), 200000, 300000, "-", "Vested: 50%"],
]
for r in other_data:
    ws.append(r)

for r in range(2, 11):
    if isinstance(ws.cell(r, 3).value, (datetime, date)):
        ws.cell(r, 3).number_format = date_fmt
    ws.cell(r, 4).number_format = currency_fmt
    ws.cell(r, 5).number_format = currency_fmt

ws["A11"] = "TOTAL"
ws["D11"] = "=SUM(D2:D10)"
ws["E11"] = "=SUM(E2:E10)"
ws["D11"].number_format = currency_fmt
ws["E11"].number_format = currency_fmt
style_total_row(ws, 11, 1, 7)
border_range(ws, 1, 11, 1, 7)

for col, w in {"A": 14, "B": 24, "C": 14, "D": 16, "E": 16, "F": 16, "G": 24}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 11: TransactionHistory ----------------
ws = wb["TransactionHistory"]
headers = [
    "Date", "Category", "Sub-Category", "Transaction Type", "Asset Name",
    "Amount (₹)", "Units/Quantity", "Price", "Account/Folio", "Remarks"
]
ws.append(headers)
style_header(ws, 1, 1, 10)

txn_data = [
    [datetime(2026, 2, 1), "Mutual Funds", "Equity", "Buy", "HDFC Flexi Cap", 10000, 11.75, 851.06, "12345/67", "SIP"],
    [datetime(2026, 2, 5), "Banking", "FD", "Maturity", "SBI FD", 210500, "-", "-", "XXXX5678", "Reinvested"],
    [datetime(2026, 2, 10), "Gold", "Physical", "Buy", "24K Coin", 55000, 10, 5500, "Bank Locker", "-"],
]
for r in txn_data:
    ws.append(r)

for r in range(2, 200):
    ws.cell(r, 1).number_format = date_fmt
    ws.cell(r, 6).number_format = currency_fmt

border_range(ws, 1, 50, 1, 10)
for col, w in {"A": 14, "B": 14, "C": 14, "D": 16, "E": 18, "F": 14, "G": 14, "H": 12, "I": 14, "J": 20}.items():
    ws.column_dimensions[col].width = w


# ---------------- Sheet 12: Reminders ----------------
ws = wb["Reminders"]
headers = ["Due Date", "Category", "Description", "Amount (₹)", "Status", "Alert Days Before"]
ws.append(headers)
style_header(ws, 1, 1, 6)

rem_data = [
    [datetime(2026, 4, 1), "Insurance", "HDFC Term Premium", 15000, "Pending", 15],
    [datetime(2026, 4, 5), "Investment", "PPF Contribution", 150000, "Pending", 30],
    [datetime(2026, 8, 15), "Banking", "SBI FD Maturity", 200000, "Pending", 30],
]
for r in rem_data:
    ws.append(r)

for r in range(2, 200):
    ws.cell(r, 1).number_format = date_fmt
    ws.cell(r, 4).number_format = currency_fmt

# Conditional formatting by due window, pending items
red_fill = PatternFill("solid", fgColor="FFC7CE")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
green_fill = PatternFill("solid", fgColor="C6EFCE")

ws.conditional_formatting.add("A2:F200", FormulaRule(formula=['=AND($E2="Pending",$A2-TODAY()<=7,$A2-TODAY()>=0)'], fill=red_fill))
ws.conditional_formatting.add("A2:F200", FormulaRule(formula=['=AND($E2="Pending",$A2-TODAY()>7,$A2-TODAY()<=15)'], fill=yellow_fill))
ws.conditional_formatting.add("A2:F200", FormulaRule(formula=['=AND($E2="Pending",$A2-TODAY()>15,$A2-TODAY()<=30)'], fill=green_fill))

border_range(ws, 1, 50, 1, 6)
for col, w in {"A": 14, "B": 14, "C": 24, "D": 14, "E": 12, "F": 16}.items():
    ws.column_dimensions[col].width = w

rem_dv = DataValidation(type="list", formula1='"Pending,Completed,Cancelled"')
ws.add_data_validation(rem_dv)
rem_dv.add("E2:E100")


# ---------------- Sheet 13: Settings ----------------
ws = wb["Settings"]
ws["A1"] = "CURRENT MARKET RATES"
ws["A1"].font = Font(bold=True, size=12)

settings_rows = [
    ("Gold (24K) per gram", 6500),
    ("Gold (22K) per gram", 5950),
    ("Silver per gram", 75),
    ("USD to INR", 83.5),
    ("Last Updated", date.today()),
]
for i, (k, v) in enumerate(settings_rows, start=2):
    ws.cell(i, 1, k)
    ws.cell(i, 2, v)

for r in [2, 3, 4]:
    ws.cell(r, 2).number_format = currency_fmt
ws.cell(5, 2).number_format = "₹#,##0.00"
ws.cell(6, 2).number_format = date_fmt

ws["A9"] = "PERSONAL SETTINGS"
ws["A9"].font = Font(bold=True, size=12)

personal_rows = [
    ("Financial Year", "2025-26"),
    ("Risk Profile", "Moderate"),
    ("Target Net Worth", 20000000),
    ("Target Date", datetime(2030, 12, 31)),
]
for i, (k, v) in enumerate(personal_rows, start=10):
    ws.cell(i, 1, k)
    ws.cell(i, 2, v)

ws["B12"].number_format = currency_fmt
ws["B13"].number_format = date_fmt

for col, w in {"A": 28, "B": 20}.items():
    ws.column_dimensions[col].width = w
border_range(ws, 1, 13, 1, 2)


# ---------------- Sheet 14: Instructions ----------------
ws = wb["Instructions"]
ws.column_dimensions["A"].width = 80
instructions_text = """INDIAN INVESTMENT PORTFOLIO TRACKER - USER GUIDE

GETTING STARTED:
1. Start by filling in your actual data in each category sheet
2. The Dashboard will automatically calculate your total net worth
3. Update the Settings sheet with current market rates

IMPORTANT FORMULAS:
• Dashboard formulas automatically pull data from all other sheets
• Do NOT delete the TOTAL rows - they contain important formulas
• Current Value columns use formulas like =Units*Price

UPDATING DATA:
• Banking: Update balances monthly
• Mutual Funds: Update NAV weekly or use live data
• Stocks: Update prices daily or use GOOGLEFINANCE function
• Gold/Silver: Update current rates from Settings sheet
• Real Estate: Update valuations annually

GOOGLE SHEETS ENHANCEMENTS:
After uploading to Google Sheets, add:
• =GOOGLEFINANCE("NSE:RELIANCE","price") for live stock prices
• =GOOGLEFINANCE("CURRENCY:XAUINR")/31.1035 for live gold rates
• Data validation dropdowns for consistent data entry
• Conditional formatting to highlight gains/losses
• Charts for visual representation

MONTHLY REVIEW CHECKLIST:
☐ Update all bank account balances
☐ Update mutual fund NAVs
☐ Update stock prices
☐ Update gold/silver rates
☐ Add new transactions to Transaction History
☐ Check Reminders for upcoming payments
☐ Review Dashboard for overall net worth

TIPS:
• Keep documents in a secure location
• Set calendar reminders for payments
• Review asset allocation quarterly
• Backup this file regularly

PRIVACY & SECURITY:
• This file contains sensitive financial information
• Store it securely and password-protect if sharing
• Do not share account numbers or passwords"""
ws["A1"] = instructions_text
ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")

# ---------------- Optional Charts on Dashboard ----------------
dash = wb["Dashboard"]

# Debt vs Equity helper
dash["A20"] = "Equity (Net Worth)"
dash["B20"] = "=B7"
dash["A21"] = "Debt (Liabilities)"
dash["B21"] = "=Loans!E11"
dash["B20"].number_format = currency_fmt
dash["B21"].number_format = currency_fmt

# Net worth trend + monthly cash flow helper
dash["J2"] = "Month"
dash["K2"] = "Net Worth"
dash["L2"] = "Income"
dash["M2"] = "Expenses"
dash["N2"] = "Investments"
for col in ["J", "K", "L", "M", "N"]:
    dash[f"{col}2"].font = Font(bold=True)

trend_months = [
    datetime(2025, 9, 1),
    datetime(2025, 10, 1),
    datetime(2025, 11, 1),
    datetime(2025, 12, 1),
    datetime(2026, 1, 1),
    datetime(2026, 2, 1),
]
trend_factors = [0.88, 0.90, 0.93, 0.96, 0.98, 1.00]

for i, (m, f) in enumerate(zip(trend_months, trend_factors), start=3):
    dash.cell(i, 10, m)  # J
    dash.cell(i, 10).number_format = "MMM-YYYY"
    dash.cell(i, 11, f"=$B$7*{f}")  # K
    dash.cell(i, 12, f'=SUMIFS(TransactionHistory!$F:$F,TransactionHistory!$A:$A,\">=\"&J{i},TransactionHistory!$A:$A,\"<\"&EDATE(J{i},1),TransactionHistory!$D:$D,\"Maturity\")+SUMIFS(TransactionHistory!$F:$F,TransactionHistory!$A:$A,\">=\"&J{i},TransactionHistory!$A:$A,\"<\"&EDATE(J{i},1),TransactionHistory!$D:$D,\"Sell\")')  # L
    dash.cell(i, 13, f'=SUMIFS(TransactionHistory!$F:$F,TransactionHistory!$A:$A,\">=\"&J{i},TransactionHistory!$A:$A,\"<\"&EDATE(J{i},1),TransactionHistory!$D:$D,\"Withdrawal\")')  # M
    dash.cell(i, 14, f'=SUMIFS(TransactionHistory!$F:$F,TransactionHistory!$A:$A,\">=\"&J{i},TransactionHistory!$A:$A,\"<\"&EDATE(J{i},1),TransactionHistory!$D:$D,\"Buy\")+SUMIFS(TransactionHistory!$F:$F,TransactionHistory!$A:$A,\">=\"&J{i},TransactionHistory!$A:$A,\"<\"&EDATE(J{i},1),TransactionHistory!$D:$D,\"Deposit\")')  # N
    for c in [11, 12, 13, 14]:
        dash.cell(i, c).number_format = currency_fmt

# Top 5 holdings helper
dash["P2"] = "Holding"
dash["Q2"] = "Value"
dash["P2"].font = Font(bold=True)
dash["Q2"].font = Font(bold=True)
top5 = [
    ("Real Estate", "=RealEstate!F11"),
    ("Market Investments", "=MutualFunds!G21+Stocks!G21"),
    ("Retirement Funds", "=Retirement!C11"),
    ("Banking & Deposits", "=Banking!D11"),
    ("Precious Metals", "=Gold!G7"),
]
for i, (name, val) in enumerate(top5, start=3):
    dash.cell(i, 16, name)  # P
    dash.cell(i, 17, val)   # Q
    dash.cell(i, 17).number_format = currency_fmt

# Hide helper columns
for col in ["J", "K", "L", "M", "N", "P", "Q"]:
    dash.column_dimensions[col].hidden = True

# 1) Asset Allocation (Pie)
pie = PieChart()
pie.title = "Asset Allocation"
pie_data = Reference(dash, min_col=2, min_row=11, max_row=17)
pie_cats = Reference(dash, min_col=1, min_row=11, max_row=17)
pie.add_data(pie_data, titles_from_data=False)
pie.set_categories(pie_cats)
pie.height = 7
pie.width = 9
dash.add_chart(pie, "F1")

# 2) Net Worth Trend (Line)
line = LineChart()
line.title = "Net Worth Trend"
line.y_axis.title = "Net Worth (₹)"
line.x_axis.title = "Month"
line_data = Reference(dash, min_col=11, min_row=2, max_row=8)
line_cats = Reference(dash, min_col=10, min_row=3, max_row=8)
line.add_data(line_data, titles_from_data=True)
line.set_categories(line_cats)
line.height = 7
line.width = 9
dash.add_chart(line, "F18")

# 3) Debt vs Equity (Doughnut)
donut = DoughnutChart()
donut.title = "Debt vs Equity"
donut_data = Reference(dash, min_col=2, min_row=20, max_row=21)
donut_cats = Reference(dash, min_col=1, min_row=20, max_row=21)
donut.add_data(donut_data, titles_from_data=False)
donut.set_categories(donut_cats)
donut.height = 7
donut.width = 9
dash.add_chart(donut, "F35")

# 4) Top 5 Holdings (Bar)
bar = BarChart()
bar.title = "Top 5 Holdings"
bar.y_axis.title = "Value (₹)"
bar.x_axis.title = "Holding"
bar_data = Reference(dash, min_col=17, min_row=2, max_row=7)
bar_cats = Reference(dash, min_col=16, min_row=3, max_row=7)
bar.add_data(bar_data, titles_from_data=True)
bar.set_categories(bar_cats)
bar.height = 7
bar.width = 9
dash.add_chart(bar, "F52")

# 5) Monthly Cash Flow (Column)
cash = BarChart()
cash.type = "col"
cash.title = "Monthly Cash Flow"
cash.y_axis.title = "Amount (₹)"
cash.x_axis.title = "Month"
cash_data = Reference(dash, min_col=12, max_col=14, min_row=2, max_row=8)
cash_cats = Reference(dash, min_col=10, min_row=3, max_row=8)
cash.add_data(cash_data, titles_from_data=True)
cash.set_categories(cash_cats)
cash.height = 7
cash.width = 9
dash.add_chart(cash, "F69")

# Freeze header rows for table-based sheets
for s in [
    "Banking", "MutualFunds", "Stocks", "Gold", "RealEstate", "Retirement",
    "Insurance", "Loans", "OtherAssets", "TransactionHistory", "Reminders"
]:
    wb[s].freeze_panes = "A2"

wb.save(OUT_FILE)
print(f"Created {OUT_FILE}")
